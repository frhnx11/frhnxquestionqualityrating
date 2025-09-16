import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from typing import List, Dict
import os
from datetime import datetime
try:
    # Try relative import first (for when running as module)
    from .ollama_analyzer import AnalysisResult
except ImportError:
    # Fall back to absolute import (for when running standalone or in PyInstaller)
    from ollama_analyzer import AnalysisResult

class ExcelGenerator:
    def __init__(self, config: Dict):
        self.output_folder = config['output']['folder']
        self.excel_filename = config['output']['excel_filename']
        self.sheet_name = config['excel']['sheet_name']
        self.auto_save_interval = config['excel']['auto_save_interval']
        
        self.workbook = None
        self.worksheet = None
        self.current_row = 2  # Start from row 2 (after headers)
        
        self.headers = [
            'Subject',
            'Topic', 
            'Subtopic',
            'Question (Complete)',
            'Answer with Explanation (Complete)',
            'Rating (out of 10)',
            'Conceptual Depth',
            'Answer Accuracy',
            'Topic-Subtopic Relevance',
            'Improved Version'
        ]
        
        self._ensure_output_folder()
        self._create_workbook()
    
    def _ensure_output_folder(self):
        """Create output folder if it doesn't exist"""
        if not os.path.exists(self.output_folder):
            os.makedirs(self.output_folder)
    
    def _create_workbook(self):
        """Create new workbook with formatted headers"""
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = self.sheet_name
        
        # Add headers
        for col, header in enumerate(self.headers, 1):
            cell = self.worksheet.cell(row=1, column=col, value=header)
            self._format_header_cell(cell)
        
        # Set column widths
        column_widths = [15, 20, 20, 60, 60, 15, 40, 40, 40, 60]
        for col, width in enumerate(column_widths, 1):
            self.worksheet.column_dimensions[get_column_letter(col)].width = width
        
        # Freeze first row
        self.worksheet.freeze_panes = 'A2'
        
        # Save initial file
        self._save_workbook()
    
    def _format_header_cell(self, cell):
        """Format header cells with styling"""
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    def _format_data_cell(self, cell, is_rating=False):
        """Format data cells with styling"""
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        if is_rating:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            # Color code ratings
            rating = cell.value
            if isinstance(rating, (int, float)):
                if rating >= 9:
                    cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Light green
                elif rating >= 7:
                    cell.fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')  # Light yellow
                elif rating >= 5:
                    cell.fill = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')  # Light orange
                else:
                    cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')  # Light red
        else:
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    def add_analysis_result(self, result: AnalysisResult) -> bool:
        """Add a single analysis result to the Excel file"""
        try:
            row_data = [
                result.subject,
                result.topic,
                result.subtopic,
                result.question_complete,
                result.answer_explanation,
                result.rating,
                result.conceptual_depth,
                result.answer_accuracy,
                result.topic_relevance,
                result.improved_version
            ]
            
            # Add data to worksheet
            for col, value in enumerate(row_data, 1):
                cell = self.worksheet.cell(row=self.current_row, column=col, value=value)
                self._format_data_cell(cell, is_rating=(col == 6))
            
            # Set row height for better readability
            self.worksheet.row_dimensions[self.current_row].height = 60
            
            self.current_row += 1
            
            # Auto-save periodically
            self._save_workbook()
            
            return True
            
        except Exception as e:
            print(f"Error adding result to Excel: {e}")
            return False
    
    def add_batch_results(self, results: List[AnalysisResult]) -> int:
        """Add multiple analysis results to Excel"""
        added_count = 0
        for result in results:
            if self.add_analysis_result(result):
                added_count += 1
        
        return added_count
    
    def _save_workbook(self):
        """Save the workbook to file"""
        try:
            filepath = os.path.join(self.output_folder, self.excel_filename)
            self.workbook.save(filepath)
            return True
        except Exception as e:
            print(f"Error saving Excel file: {e}")
            return False
    
    def finalize_excel(self):
        """Add final formatting and summary statistics"""
        try:
            # Add summary row
            summary_row = self.current_row + 2
            
            # Add summary header
            summary_cell = self.worksheet.cell(row=summary_row, column=1, value="ANALYSIS SUMMARY")
            summary_cell.font = Font(bold=True, size=14)
            summary_cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            
            # Merge cells for summary header
            self.worksheet.merge_cells(f'A{summary_row}:D{summary_row}')
            
            # Calculate statistics
            if self.current_row > 2:
                total_questions = self.current_row - 2
                
                # Average rating
                avg_rating_cell = self.worksheet.cell(row=summary_row + 1, column=1, value="Average Rating:")
                avg_rating_value = self.worksheet.cell(row=summary_row + 1, column=2, 
                                                     value=f"=AVERAGE(F2:F{self.current_row - 1})")
                
                # Count by rating ranges
                high_count = self.worksheet.cell(row=summary_row + 2, column=1, value="High Quality (9-10):")
                high_count_value = self.worksheet.cell(row=summary_row + 2, column=2,
                                                      value=f'=COUNTIFS(F2:F{self.current_row - 1},">=9")')
                
                good_count = self.worksheet.cell(row=summary_row + 3, column=1, value="Good Quality (7-8):")
                good_count_value = self.worksheet.cell(row=summary_row + 3, column=2,
                                                      value=f'=COUNTIFS(F2:F{self.current_row - 1},">=7",F2:F{self.current_row - 1},"<9")')
                
                avg_count = self.worksheet.cell(row=summary_row + 4, column=1, value="Average Quality (5-6):")
                avg_count_value = self.worksheet.cell(row=summary_row + 4, column=2,
                                                     value=f'=COUNTIFS(F2:F{self.current_row - 1},">=5",F2:F{self.current_row - 1},"<7")')
                
                low_count = self.worksheet.cell(row=summary_row + 5, column=1, value="Below Standard (<5):")
                low_count_value = self.worksheet.cell(row=summary_row + 5, column=2,
                                                     value=f'=COUNTIF(F2:F{self.current_row - 1},"<5")')
                
                # Total count
                total_cell = self.worksheet.cell(row=summary_row + 6, column=1, value="Total Questions:")
                total_value = self.worksheet.cell(row=summary_row + 6, column=2, value=total_questions)
                
                # Format summary cells
                for cell in [avg_rating_cell, high_count, good_count, avg_count, low_count, total_cell]:
                    cell.font = Font(bold=True)
            
            # Add metadata
            metadata_row = summary_row + 8
            metadata_cell = self.worksheet.cell(row=metadata_row, column=1, 
                                               value=f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            metadata_cell.font = Font(italic=True, size=10)
            
            # Final save
            self._save_workbook()
            
        except Exception as e:
            print(f"Error finalizing Excel: {e}")
    
    def get_excel_path(self) -> str:
        """Get the full path to the Excel file"""
        return os.path.join(self.output_folder, self.excel_filename)
    
    def get_progress_info(self) -> Dict:
        """Get current progress information"""
        return {
            'total_processed': max(0, self.current_row - 2),
            'excel_path': self.get_excel_path(),
            'last_updated': datetime.now().strftime('%H:%M:%S')
        }